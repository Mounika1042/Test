from openpyxl import load_workbook

# Function to process the uploaded Excel file
def process_excel_file(file_path):
    try:
        # Load the Excel workbook
        workbook = load_workbook(filename=file_path, data_only=True)

        # Access a specific worksheet (e.g., the first one)
        worksheet = workbook.active

        # Process the data (example: print the contents)
        for row in worksheet.iter_rows():
            for cell in row:
                print(cell.value, end='\t')
            print()

        print("Excel file processed successfully.")

    except Exception as e:
        print("Error processing the Excel file:", str(e))

if __name__ == "__main__":
    file_path = input("Enter the path to the Excel file: ")
    process_excel_file(file_path)
